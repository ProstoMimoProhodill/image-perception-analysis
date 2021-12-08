from openpyxl import load_workbook

columns_BASIC = {
    'Procedure': 'AA',
    'Mask1': 'BS',
    'Mask2': 'BT',
    'Mask3': 'BU',
    'Mask4': 'BV',
    'Mask5': 'BW',
    'Mask6': 'BX',
    'Stimulus1': 'CA',
    'Stimulus1.ACC': 'CB',
    'Stimulus1.RT': 'CE',
    'Stimulus2': 'CG',
    'Stimulus2.ACC': 'CH',
    'Stimulus2.RT': 'CK',
    'Stimulus3': 'CM',
    'Stimulus3.ACC': 'CN',
    'Stimulus3.RT': 'CQ',
    'Stimulus4': 'CS',
    'Stimulus4.ACC': 'CT',
    'Stimulus4.RT': 'CW',
    'Stimulus5': 'CY',
    'Stimulus5.ACC': 'CZ',
    'Stimulus5.RT': 'DC',
    'Stimulus6': 'DE',
    'Stimulus6.ACC': 'DF',
    'Stimulus6.RT': 'DI',
}

columns_SUPER = {
    'Procedure': 'AA',
    'Mask1': 'BQ',
    'Mask2': 'BR',
    'Mask3': 'BS',
    'Mask4': 'BT',
    'Mask5': 'BU',
    'Mask6': 'BV',
    'Stimulus1': 'BY',
    'Stimulus1.ACC': 'BZ',
    'Stimulus1.RT': 'CC',
    'Stimulus2': 'CE',
    'Stimulus2.ACC': 'CF',
    'Stimulus2.RT': 'CI',
    'Stimulus3': 'CK',
    'Stimulus3.ACC': 'CL',
    'Stimulus3.RT': 'CO',
    'Stimulus4': 'CQ',
    'Stimulus4.ACC': 'CR',
    'Stimulus4.RT': 'CU',
    'Stimulus5': 'CW',
    'Stimulus5.ACC': 'CX',
    'Stimulus5.RT': 'DA',
    'Stimulus6': 'DC',
    'Stimulus6.ACC': 'DD',
    'Stimulus6.RT': 'DG',
}

procedures = {
    'First': '1',
    'Second': '2',
    'Third': '3',
    'Fourth': '4',
    'Fifth': '5',
    'Sixth': '6',
    'TrainingPr': None
}

def category_BASIC(m, s):
    third = ['baraban', 'had', 'key', 'kolokolchik', 'kuvshin', 'lock', 'vedro', 'ventilator']
    if m == s:
        return 1
    elif m in third:
        return 3
    else:
        return 2

def category_SUPER(m, s):
    first = ['divan_1.bmp', 'divan_2.bmp', 'divan_101.bmp', 'divan_102.bmp', 'krovat_1.bmp', 'krovat_2.bmp', 'puf.bmp', 'Puf_101.bmp', 'chair_1.bmp', 'chair_2.bmp', 'kom_101.bmp', 'kreslo1.bmp', 'kreslo2.bmp', 'table_1.bmp', 'table_2.bmp', 'tumba1.bmp']
    second = ['kastrulia_1.bmp', 'kastrulia_2.bmp', 'kastrulia_103.bmp', 'kastrulia_104.bmp', 'molochknik_1.bmp', 'saharnica_3.bmp', 'sousnik_1.bmp', 'saharnica_2.bmp', 'chainik_1.bmp', 'chainik_2.bmp', 'chainik_5.bmp', 'chainik_6.bmp', 'cup_1.bmp', 'cup_2.bmp', 'pan_101.bmp', 'pan_102.bmp']
    third = ['binokl_1.bmp', 'botinok_3.bmp', 'iron_2.bmp', 'kleczhi_1.bmp', 'kolyaska_1.bmp', 'venik_1.bmp', 'violin_2.bmp', 'zont_1.bmp']
    if m in third:
        return 3
    elif m in first and s in first or m in second and s in second:
        return 1
    else:
        return 2


def extract_data(filename):
    res = []
    ws = load_workbook(filename).active
    data_type = filename[9:-5].split('_')[-1]
    if data_type == 'Basic':
        for i in range(2, len(ws[columns_BASIC['Procedure']])):
            procedure = procedures[ws[columns_BASIC['Procedure']][i].value]
            if procedure:
                stimulus_ACC = ws[columns_BASIC['Stimulus'+procedure+'.ACC']][i].value
                if int(stimulus_ACC):
                    mask = ws[columns_BASIC['Mask'+procedure]][i].value
                    stimulus = ws[columns_BASIC['Stimulus'+procedure]][i].value
                    stimulus_RT = ws[columns_BASIC['Stimulus'+procedure+'.RT']][i].value
                    mask_category = mask.split('_')[0]
                    stimulus_category = stimulus.split('_')[0]
                    category = category_BASIC(mask_category, stimulus_category)
                    res.append((procedure, mask, stimulus, stimulus_ACC, stimulus_RT, category))
    elif data_type == 'Super':
        for i in range(2, len(ws[columns_SUPER['Procedure']])):
            procedure = procedures[ws[columns_SUPER['Procedure']][i].value]
            if procedure:
                stimulus_ACC = ws[columns_SUPER['Stimulus'+procedure+'.ACC']][i].value
                if int(stimulus_ACC):
                    mask = ws[columns_SUPER['Mask'+procedure]][i].value
                    stimulus = ws[columns_SUPER['Stimulus'+procedure]][i].value
                    stimulus_RT = ws[columns_SUPER['Stimulus'+procedure+'.RT']][i].value
                    category = category_SUPER(mask, stimulus)
                    res.append((procedure, mask, stimulus, stimulus_ACC, stimulus_RT, category))
    return res