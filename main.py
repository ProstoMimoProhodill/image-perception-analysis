import glob
import openpyxl
from util import *

files = glob.glob('raw-data/*.xlsx')
for filename in files:
    print("START {}".format(filename))
    done_filename = 'done-data/DONE_{}'.format(filename[9:])
    if glob.glob(done_filename):
        print("EXIST {}".format(filename))
        continue

    data = extract_data(filename)
    wb = openpyxl.Workbook()
    ws = wb.active
    top = ['Procedure', 'Mask', 'Stimulus', 'Stimulus.ACC', 'Stimulus.RT', 'Categories']
    COLUMN = 1
    for p in range(6):
        RT_1, RT_1_C, RT_2, RT_2_C, RT_3, RT_3_C = 0, 0, 0, 0, 0, 0
        ROW = 0
        for i in range(len(top)):
            cell = ws.cell(row=1, column=COLUMN+i, value=top[i])
        for i in range(len(data)):
            if data[i][0] == str(p+1):
                if data[i][5] == 1:
                    RT_1_C += 1
                    RT_1 += data[i][4]
                elif data[i][5] == 2:
                    RT_2_C += 1
                    RT_2 += data[i][4]
                elif data[i][5] == 3:
                    RT_3_C += 1
                    RT_3 += data[i][4]
                for j in range(len(data[i])):
                    cell = ws.cell(row=2+ROW, column=COLUMN+j, value=data[i][j])
                ROW += 1
        # AVERAGE
        ROW = 100
        cell = ws.cell(row=ROW, column=COLUMN, value='Average')
        for i in range(1, 4):
            avg = 0
            if i == 1:
                avg = RT_1 / RT_1_C
            elif i == 2:
                avg = RT_2 / RT_2_C
            elif i == 3:
                avg = RT_3 / RT_3_C
            cell = ws.cell(row=ROW, column=COLUMN+i+1, value=avg)
        # COUNT
        ROW = 102
        cell = ws.cell(row=ROW, column=COLUMN, value='Count')
        for i in range(1, 4):
            c = 0
            if i == 1:
                c = RT_1_C
            elif i == 2:
                c = RT_2_C
            elif i == 3:
                c = RT_3_C
            cell = ws.cell(row=ROW, column=COLUMN+i+1, value=c)
        # PERCENT
        ROW = 104
        cell = ws.cell(row=ROW, column=COLUMN, value='Relation')
        for i in range(1, 4):
            c = 0
            if i == 1:
                c = RT_1_C / 32
            elif i == 2:
                c = RT_2_C / 32
            elif i == 3:
                c = RT_3_C / 32
            cell = ws.cell(row=ROW, column=COLUMN+i+1, value=c)
        COLUMN += 9
        wb.save(done_filename)

    print("SAVE {}".format(filename))
